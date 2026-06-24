from utils.utils import *
from utils.constants import *
from utils.model_player import *
from utils.player_constants import *
from catanatron.game import Game, TURNS_LIMIT
from itertools import combinations
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PLAYOUT_GAMES = 500

colors = [Color.BLUE, Color.RED]

env = create_random_players_env(reward_function=None, num_enemies=1)
observation, _ = env.reset()
env.close()

OBSERVATION_SHAPE = observation.shape[0]

def create_model_player(model_stats: dict, color: Color) -> ModelPlayer:
    model = model_stats["model_type"](OBSERVATION_SHAPE, MAX_ACTION_COUNT, neurons = model_stats["neurons"]).to(device)
    model.load_state_dict(torch.load(os.path.join(MODELS_SAVE_PATH, model_stats["save_path"]), map_location=device))
    model.eval()
    player = ModelPlayer(model, color, is_bot = True, device = device)
    return player

# Let the games begin
# leave empty in case you want to start evaluation again. Those are already made evaluations so we don't always start over again.
results = [
    ["D3QN (Dueling, Double)", "D2QN Big Selfplay (Dueling)", 486, 514],
    ["D3QN (Dueling, Double)", "D2QN Small Selfplay (Dueling)", 553, 447],
    ["D3QN (Dueling, Double)", "D2QN Small (Dueling)", 427, 573],
    ["D3QN (Dueling, Double)", "D2QN Big (Dueling)", 621, 379],
    ["D3QN (Dueling, Double)", "PPO DefaultR Big", 732, 268],
    ["D3QN (Dueling, Double)", "PPO BetterR Big", 561, 439],
    ["D3QN (Dueling, Double)", "PPO BetterR Small", 376, 624],
    ["D3QN (Dueling, Double)", "PPO DefaultR Big", 734, 266],
    ["D3QN (Dueling, Double)", "REINFORCE", 993, 7],

    ["D2QN Big Selfplay (Dueling)", "D2QN Small Selfplay (Dueling)", 483, 517],
    ["D2QN Big Selfplay (Dueling)", "D2QN Small (Dueling)", 259, 741],
    ["D2QN Big Selfplay (Dueling)", "D2QN Big (Dueling)", 629, 371],
    ["D2QN Big Selfplay (Dueling)", "PPO DefaultR Big", 801, 199],
    ["D2QN Big Selfplay (Dueling)", "PPO BetterR Big", 582, 418],
    ["D2QN Big Selfplay (Dueling)", "PPO BetterR Small", 814, 186],
    ["D2QN Big Selfplay (Dueling)", "REINFORCE", 497, 3],

    ["D2QN Small Selfplay (Dueling)", "D2QN Small (Dueling)", 296, 204],
    ["D2QN Small Selfplay (Dueling)", "D2QN Big (Dueling)", 307, 193],
    ["D2QN Small Selfplay (Dueling)", "PPO DefaultR Big", 351, 149],
    ["D2QN Small Selfplay (Dueling)", "PPO BetterR Big", 258, 242],
    ["D2QN Small Selfplay (Dueling)", "PPO BetterR Small", 328, 172],
    ["D2QN Small Selfplay (Dueling)", "REINFORCE", 478, 22],

    ["D2QN Small (Dueling)", "D2QN Big (Dueling)", 149, 351],
    ["D2QN Small (Dueling)", "PPO DefaultR Big", 339, 161],
    ["D2QN Small (Dueling)", "PPO BetterR Big", 222, 278],
    ["D2QN Small (Dueling)", "PPO BetterR Small", 248, 252],
    ["D2QN Small (Dueling)", "REINFORCE", 498, 2],

    ["D2QN Big (Dueling)", "PPO DefaultR Big", 373, 127],
    ["D2QN Big (Dueling)", "PPO BetterR Big", 248, 252],
    ["D2QN Big (Dueling)", "PPO BetterR Small", 299, 201],
    ["D2QN Big (Dueling)", "REINFORCE", 495, 5],

    ["PPO DefaultR Big", "PPO BetterR Big", 207, 293],
    ["PPO DefaultR Big", "PPO BetterR Small", 183, 317],
    ["PPO DefaultR Big", "REINFORCE", 486, 14],

    ["PPO BetterR Big", "PPO BetterR Small", 227, 273],
    ["PPO BetterR Big", "REINFORCE", 499, 1],

    ["PPO BetterR Small", "REINFORCE", 484, 16],

    ["D3QN (Dueling, Double)", "RandomPlayer", 436, 64],
    ["D3QN (Dueling, Double)", "Weighted Random Player", 405, 95],
    ["D3QN (Dueling, Double)", "Victory Point Player", 388, 112],
    ["D3QN (Dueling, Double)", "Alpha Beta Player", 5, 495],

    ["D3QN (Dueling, Double)", "MCTS Player", 470, 30],
    ["D3QN (Dueling, Double)", "Value Function Player", 10, 490],

    ["D2QN Big Selfplay (Dueling)", "RandomPlayer", 472, 28],
    ["D2QN Big Selfplay (Dueling)", "Weighted Random Player", 450, 50],
    ["D2QN Big Selfplay (Dueling)", "Victory Point Player", 413, 87],
    ["D2QN Big Selfplay (Dueling)", "Alpha Beta Player", 9, 491],
    ["D2QN Big Selfplay (Dueling)", "MCTS Player", 454, 46],
    ["D2QN Big Selfplay (Dueling)", "Value Function Player", 25, 475],

    ["D2QN Small Selfplay (Dueling)", "RandomPlayer", 450, 50],
    ["D2QN Small Selfplay (Dueling)", "Weighted Random Player", 448, 52],
    ["D2QN Small Selfplay (Dueling)", "Victory Point Player", 404, 96],
    ["D2QN Small Selfplay (Dueling)", "Alpha Beta Player", 21, 479],
    ["D2QN Small Selfplay (Dueling)", "MCTS Player", 460, 40],
    ["D2QN Small Selfplay (Dueling)", "Value Function Player", 20, 480],
    
    ["D2QN Small (Dueling)", "RandomPlayer", 464, 36],
    ["D2QN Small (Dueling)", "Weighted Random Player", 422, 78],
    ["D2QN Small (Dueling)", "Victory Point Player", 453, 47],
    ["D2QN Small (Dueling)", "Alpha Beta Player", 25, 475],
    ["D2QN Small (Dueling)", "MCTS Player", 448, 52],
    ["D2QN Small (Dueling)", "Value Function Player", 22, 478],

    ["D2QN Big (Dueling)", "RandomPlayer", 468, 32],
    ["D2QN Big (Dueling)", "Weighted Random Player", 441, 59],
    ["D2QN Big (Dueling)", "Victory Point Player", 423, 77],
    ["D2QN Big (Dueling)", "Alpha Beta Player", 17, 483],
    ["D2QN Big (Dueling)", "MCTS Player", 437, 63],
    ["D2QN Big (Dueling)", "Value Function Player", 18, 482],

    ["PPO DefaultR Big", "RandomPlayer", 382, 118],
    ["PPO DefaultR Big", "Weighted Random Player", 347, 153],
    ["PPO DefaultR Big", "Victory Point Player", 312, 188],
    ["PPO DefaultR Big", "Alpha Beta Player", 31, 469],
    ["PPO DefaultR Big", "MCTS Player", 363, 137],
    ["PPO DefaultR Big", "Value Function Player", 53, 447],

    ["PPO BetterR Big", "RandomPlayer", 418, 82],
    ["PPO BetterR Big", "Weighted Random Player", 391, 109],
    ["PPO BetterR Big", "Victory Point Player", 353, 147],
    ["PPO BetterR Big", "Alpha Beta Player", 22, 478],
    ["PPO BetterR Big", "MCTS Player", 403, 97],
    ["PPO BetterR Big", "Value Function Player", 41, 459],

    ["PPO BetterR Small", "RandomPlayer", 358, 142],
    ["PPO BetterR Small", "Weighted Random Player", 323, 177],
    ["PPO BetterR Small", "Victory Point Player", 287, 213],
    ["PPO BetterR Small", "Alpha Beta Player", 38, 462],
    ["PPO BetterR Small", "MCTS Player", 341, 159],
    ["PPO BetterR Small", "Value Function Player", 62, 438],

    ["REINFORCE", "RandomPlayer", 176, 324],
    ["REINFORCE", "Weighted Random Player", 152, 348],
    ["REINFORCE", "Victory Point Player", 129, 371],
    ["REINFORCE", "Alpha Beta Player", 4, 496],
    ["REINFORCE", "MCTS Player", 163, 337],
    ["REINFORCE", "Value Function Player", 11, 489],
]

def check_repetition(name1, name2) -> bool:
    for result in results:
        if (result[0] == name1 or result[1] == name1) and (result[0] == name2 or result[1] == name2):
            return True
    return False

def play_games(player1: Player, player2: Player) -> tuple[int, int]:
    player1_wins = 0
    player2_wins = 0
    for i in range(PLAYOUT_GAMES):
        game: Game = Game(players = [player1, player2])
        game.play()
        while game.state.num_turns == TURNS_LIMIT:
            game: Game = Game(players = [player1, player2])
            game.play()
            print(f"Repeating game {i}")
        if game.winning_color() == player1.color:
            player1_wins += 1
        else:
            player2_wins += 1
    return (player1_wins, player2_wins)

for (name1, name2) in combinations(model_players, 2):
    if check_repetition(name1, name2):
        continue
    player1: ModelPlayer = create_model_player(model_players[name1], Color.BLUE)
    player2: ModelPlayer = create_model_player(model_players[name2], Color.RED)
    
    player1_wins, player2_wins = play_games(player1, player2)
    print(f"{name1} vs {name2} - {player1_wins} : {player2_wins}")

    # TODO!!: this is a bit ugly.
    # TODO: Also, should I log more information?
    results.append([name1, name2, player1_wins, player2_wins])


# TODO: some function to combine, not two fors?
for model_player_name in model_players:
    for cpu_player_name in cpu_players:
        if check_repetition(model_player_name, cpu_player_name):
            continue
        player1: ModelPlayer = create_model_player(model_players[model_player_name], Color.BLUE)
        player2: ModelPlayer = cpu_players[cpu_player_name](Color.RED)

        player1_wins, player2_wins = play_games(player1, player2)
        print(f"{model_player_name} vs {cpu_player_name} - {player1_wins} : {player2_wins}")

        results.append([model_player_name, cpu_player_name, player1_wins, player2_wins])

# no time for this
# for (name1, name2) in combinations(cpu_players, 2):
#     if check_repetition(name1, name2):
#         continue
#     player1: ModelPlayer = cpu_players[name1](Color.BLUE)
#     player2: ModelPlayer = cpu_players[name2](Color.RED)

#     player1_wins, player2_wins = play_games(player1, player2)
#     print(f"{name1} vs {name2} - {player1_wins} : {player2_wins}")

#     results.append([name1, name2, player1_wins, player2_wins])

# The code below is AI generated----------------------------------------------

def save_results_to_excel(results, filename="model_evaluation.xlsx"):
    all_model_names = list(model_players.keys())
    all_cpu_names   = list(cpu_players.keys())
    all_names       = all_model_names + all_cpu_names

    matchup = {}
    for name1, name2, w1, w2 in results:
        matchup[(name1, name2)] = (w1, w2)
        matchup[(name2, name1)] = (w2, w1)

    total_wins  = defaultdict(int)
    total_games = defaultdict(int)
    for name1, name2, w1, w2 in results:
        total_wins[name1]  += w1;  total_wins[name2]  += w2
        total_games[name1] += w1 + w2;  total_games[name2] += w1 + w2

    MODEL_FILL  = PatternFill("solid", fgColor="FFE699")  # gold
    HEADER_FILL = PatternFill("solid", fgColor="2F5496")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    bold_font   = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    for col, h in enumerate(["Player", "Type", "Wins", "Losses", "Total Games", "Win %", "Comment"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL;  c.font = header_font;  c.alignment = center

    for row, name in enumerate(all_names, 2):
        is_model = name in model_players
        fill     = MODEL_FILL if is_model else None
        w = total_wins[name];  g = total_games[name]
        comment  = model_players[name]["comment"] if is_model else ""
        row_data = [name, "Model" if is_model else "CPU", w, g - w, g,
                    round(w / g * 100, 1) if g else 0.0, comment]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            if fill: c.fill = fill

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 45

    # ── Sheet 2: Matchup Matrix ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Matchup Matrix")

    ws2.cell(row=1, column=1, value="vs →").font = bold_font
    for col, name in enumerate(all_names, 2):
        c = ws2.cell(row=1, column=col, value=name)
        c.fill = MODEL_FILL if name in model_players else HEADER_FILL
        c.font = Font(bold=True) if name in model_players else header_font
        c.alignment = Alignment(horizontal="center", textRotation=45)

    for row, row_name in enumerate(all_names, 2):
        is_model_row = row_name in model_players
        c = ws2.cell(row=row, column=1, value=row_name)
        c.font = bold_font
        if is_model_row: c.fill = MODEL_FILL

        for col, col_name in enumerate(all_names, 2):
            if row_name == col_name:
                ws2.cell(row=row, column=col, value="—").alignment = center
                continue
            key = (row_name, col_name)
            if key in matchup:
                w, l = matchup[key]
                total = w + l
                pct = round(w / total * 100, 1) if total else 0.0
                c = ws2.cell(row=row, column=col, value=f"{pct}%\n({w}-{l})")
                c.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                if is_model_row: c.fill = MODEL_FILL
            else:
                ws2.cell(row=row, column=col, value="N/A").alignment = center

    ws2.column_dimensions["A"].width = 35
    ws2.row_dimensions[1].height = 90
    for col in range(2, len(all_names) + 2):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    path = os.path.join(STATS_SAVE_PATH, filename)
    wb.save(path)
    print(f"Saved evaluation to {path}")


save_results_to_excel(results)