"""
Run this script from the src/ directory to generate a fake model_evaluation.xlsx
for testing the excel layout without playing any actual games.
"""

import os
import random
from itertools import combinations
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

DOCUMENTATION_SAVE_PATH = "../documentation/"
PLAYOUT_GAMES = 1000

model_players = {
    "D3QN (Dueling, Double)": {
        "comment": "Dueling + Double DQN, 512 neurons. Trained 12K episodes vs WeightedRandom with shaped reward.",
    },
    "D2QN Big Selfplay (Dueling)": {
        "comment": "Dueling DQN, 1024 neurons. Trained 15K episodes via self-play (opponent synced every 100 episodes).",
    },
    "D2QN Small Selfplay (Dueling)": {
        "comment": "Dueling DQN, 512 neurons. Trained 9K episodes via self-play (opponent synced every 100 episodes).",
    },
    "D2QN Small (Dueling)": {
        "comment": "Dueling DQN, 512 neurons. Trained 15K episodes vs WeightedRandom with shaped reward.",
    },
    "D2QN Big (Dueling)": {
        "comment": "Dueling DQN, 1024 neurons. Only 4K episodes — undertrained vs WeightedRandom, needs more training.",
    },
    "PPO DefaultR Big": {
        "comment": "PPO Actor, 1024 neurons. Trained 10K episodes with default catanatron reward vs WeightedRandom. Needs more training.",
    },
    "PPO BetterR Small": {
        "comment": "PPO Actor, 512 neurons. Trained 15K episodes with shaped reward vs WeightedRandom.",
    },
    "REINFORCE": {
        "comment": "Vanilla REINFORCE (Monte Carlo policy gradient), 512 neurons. Trained 16K episodes with shaped reward vs WeightedRandom.",
    },
}

cpu_players = {
    "Weighted Random Player": None,
    "Victory Point Player": None,
    "Alpha Beta Player": None,
    "Same Turn Alpha Beta Player": None,
    "Greedy Playouts Player": None,
    "MCTS Player": None,
    "Value Function Player": None,
}


def fake_wins(total=PLAYOUT_GAMES):
    w = random.randint(0, total)
    return w, total - w


results = []

for name1, name2 in combinations(model_players, 2):
    w1, w2 = fake_wins()
    results.append([name1, name2, w1, w2])

for model_name in model_players:
    for cpu_name in cpu_players:
        w1, w2 = fake_wins()
        results.append([model_name, cpu_name, w1, w2])

for name1, name2 in combinations(cpu_players, 2):
    w1, w2 = fake_wins()
    results.append([name1, name2, w1, w2])


def save_results_to_excel(results, filename="model_evaluation.xlsx"):
    all_model_names = list(model_players.keys())
    all_cpu_names   = list(cpu_players.keys())
    all_names       = all_model_names + all_cpu_names

    matchup = {}
    for name1, name2, w1, w2 in results:
        matchup[(name1, name2)] = (w1, w2)
        matchup[(name2, name1)] = (w2, w1)

    total_wins  = defaultdict(int)
    total_games = defaultdict(int)
    for name1, name2, w1, w2 in results:
        total_wins[name1]  += w1;  total_wins[name2]  += w2
        total_games[name1] += w1 + w2;  total_games[name2] += w1 + w2

    MODEL_FILL  = PatternFill("solid", fgColor="FFE699")
    HEADER_FILL = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font   = Font(bold=True)
    center      = Alignment(horizontal="center", vertical="center")

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"

    for col, h in enumerate(["Player", "Type", "Wins", "Losses", "Total Games", "Win %", "Comment"], 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL;  c.font = header_font;  c.alignment = center

    for row, name in enumerate(all_names, 2):
        is_model = name in model_players
        fill     = MODEL_FILL if is_model else None
        w = total_wins[name];  g = total_games[name]
        comment  = model_players[name]["comment"] if is_model else ""
        row_data = [name, "Model" if is_model else "CPU", w, g - w, g,
                    round(w / g * 100, 1) if g else 0.0, comment]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            if fill: c.fill = fill

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 45

    ws2 = wb.create_sheet("Matchup Matrix")

    ws2.cell(row=1, column=1, value="vs →").font = bold_font
    for col, name in enumerate(all_names, 2):
        c = ws2.cell(row=1, column=col, value=name)
        c.fill = MODEL_FILL if name in model_players else HEADER_FILL
        c.font = Font(bold=True) if name in model_players else header_font
        c.alignment = Alignment(horizontal="center", textRotation=45)

    for row, row_name in enumerate(all_names, 2):
        is_model_row = row_name in model_players
        c = ws2.cell(row=row, column=1, value=row_name)
        c.font = bold_font
        if is_model_row: c.fill = MODEL_FILL

        for col, col_name in enumerate(all_names, 2):
            if row_name == col_name:
                ws2.cell(row=row, column=col, value="—").alignment = center
                continue
            key = (row_name, col_name)
            if key in matchup:
                w, l = matchup[key]
                total = w + l
                pct = round(w / total * 100, 1) if total else 0.0
                c = ws2.cell(row=row, column=col, value=f"{pct}%\n({w}-{l})")
                c.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
                if is_model_row: c.fill = MODEL_FILL
            else:
                ws2.cell(row=row, column=col, value="N/A").alignment = center

    ws2.column_dimensions["A"].width = 35
    ws2.row_dimensions[1].height = 90
    for col in range(2, len(all_names) + 2):
        ws2.column_dimensions[get_column_letter(col)].width = 14

    path = os.path.join(DOCUMENTATION_SAVE_PATH, filename)
    wb.save(path)
    print(f"Saved evaluation to {path}")


save_results_to_excel(results, filename="fake_model_evaluation.xlsx")
